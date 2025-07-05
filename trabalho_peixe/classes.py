import qdarkstyle
from time import sleep
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from PySide6.QtWidgets import (QMainWindow, QLineEdit, QWidget, QVBoxLayout,
                               QGridLayout, QLabel, QPushButton, QMessageBox, QMenuBar)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QKeyEvent
from variaveis import *
from funcoes import isNumero, isLetra
from openpyxl import load_workbook
from pathlib import Path


EXCEL = 'arquivo.xlsx'

PASTA_MAE = Path(__file__).parent
PASTA_EXCEL = PASTA_MAE / 'arquivo_excel'
ARQUIVO = PASTA_EXCEL / EXCEL

class Planilha(Workbook):
    def __init__(self):
        super().__init__()
        self.worksheet = Worksheet
        self.ativo = self.active
        self.outroCelula()
        self._repetir(ARQUIVO)

        
    def outroCelula(self, nome : str | None = None):
        del self['Sheet']
        if nome is not None:
            self.create_sheet(f'{nome}', 0)
            self.worksheet: Worksheet = self[f'{nome}']
            self.worksheet.cell(1, 1, 'Nome')
            self.worksheet.cell(1, 2, 'Kg')
            return
        file = 'MinhaPlanilha'
        self.create_sheet(file)
        self.worksheet: Worksheet = self[f'{file}']
        self.worksheet.cell(1, 1, 'Nome')
        self.worksheet.cell(1, 2, 'Kg')
        

    def _repetir(self, caminho):
        try:
            self.load = load_workbook(caminho)
            wac = self.load.active

            lista = []
            listas = []
            for row in wac.iter_rows():
                for cell in row:
                    lista.append(str(cell.value))
                listas.append(lista.copy())
                lista.clear()

            self._colocar(listas)
        except FileNotFoundError:
            None

    def _colocar(self, lista: list):
        for i ,listar in enumerate(lista):
            if i == 0:
                continue
            self.worksheet.append(listar)
            
    def salvar(self):
        self.save(ARQUIVO)


class MainWindow(QMainWindow):
    def __init__(self, parent: QWidget| None=None, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.widget = QWidget()
        self.vLayout = QVBoxLayout()

        self.widget.setLayout(self.vLayout)
        self.setCentralWidget(self.widget)

    def adjustFixedSize(self):
        self.adjustSize()
        self.setFixedSize(self.width(), self.height())

    def addObjetoLayout(self, widgets: QWidget):
        self.vLayout.addWidget(widgets)
    
    def Fazercaixa(self):
        return QMessageBox(self)


class Usuario(QLineEdit):
    eqPressed = Signal(str)

    def __init__(self, *args, **kwagrs):
        super().__init__(*args, **kwagrs)
        self._confyStyle()

    def _confyStyle(self):
        self.setStyleSheet(f'font-size: {BIG_FONT_SIZE}PX')
        self.setMinimumHeight(BIG_FONT_SIZE)
        self.setMinimumWidth(MINIMUM_WIDTH) 
        self.setAlignment(Qt.AlignmentFlag.AlignRight)
        self.setTextMargins(*[TEXT_MARGIN - 10 for _ in range(4)])

class Valor(QLineEdit):
    def __init__(self, *args, **kwagrs):
        super().__init__(*args, **kwagrs)
        self._confyStyle()

    def _confyStyle(self):
        self.setStyleSheet(f'font-size: {BIG_FONT_SIZE}PX')
        self.setMinimumHeight(BIG_FONT_SIZE)
        self.setMinimumWidth(MINIMUM_WIDTH) 
        self.setAlignment(Qt.AlignmentFlag.AlignRight)
        self.setTextMargins(*[TEXT_MARGIN - 10 for _ in range(4)])
    

class Info(QLabel):
    def __init__(self, text: str, parents: QWidget | None= None):
        super().__init__(text ,parents)
        self._configurarStilo()

    def _configurarStilo(self):
        self.setStyleSheet(f'font-size: {SMALL_FONT_SIZE}px')
        self.setAlignment(Qt.AlignmentFlag.AlignCenter)
        # self.setAlignment(Qt.AlignmentFlag.AlignTop)

class Button(QPushButton):
    def __init__(self, *args, **Kwargs):
        super().__init__(*args, **Kwargs)
        self.config()

    def config(self):
        fonte = self.font()
        fonte.setPixelSize(SMALL_FONT_SIZE)
        self.setFont(fonte)
        self.setMinimumSize(75, 75)


class Grades(QGridLayout):
    def __init__(self, planilha: Planilha ,usuario: Usuario ,valor: Valor,info: Info ,*args, **kwagrs):
        super().__init__(*args, **kwagrs)
        self.planilha = planilha
        self.usuario = usuario
        self.valor = valor
        self.info = info
        self.box = MainWindow()

        self._salvar()

    def _salvar(self):
        button = Button('Salvar')
        self.addWidget(button)
        slot = self._soltCriar(self._criarAcao)
        button.clicked.connect(slot)
    
    def _soltCriar(self, func, *args, **kwargs):
        def atraso():
            func(*args, **kwargs)
        return atraso

    # def _pressionar(self):
    #     self.usuario.eqPressed.connect()
    
    def _inserirText(self):
        if self.usuario.text() == '':
            return self._faltando('nome do cliente:) ')
        
        elif self.valor.text() == '':
            return self._faltando('Valor invalido')

    def _faltando(self, texto):
        text = self.box.Fazercaixa()
        text.setText(texto)
        text.exec()

    def _criarAcao(self):
        self._inserirText()

        if isNumero(self.valor.text()) and isLetra(self.usuario.text()):
            self.planilha.worksheet.append([self.usuario.text(), self.valor.text()])
        else:
            self._faltando('dados incorretos')
        
        self.usuario.clear()
        self.valor.clear()
    

css = """
    QWidgets {
        background-color: #26486B;
        color: #60798B ;
    }
"""

def setupTheme(app):
    app.setStyleSheet(qdarkstyle.load_stylesheet())
    app.setStyleSheet(app.styleSheet() + css)